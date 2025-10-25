#!/usr/bin/env python3
"""
Excel Workbook Analyzer - Extract full Excel content locally with zero LLM involvement
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import json
import sys
from pathlib import Path
from datetime import datetime, time, date

# Import shared smart cache
sys.path.insert(0, str(Path(__file__).parent.parent.parent / 'shared'))
from smart_cache import SmartCache

def serialize_cell_value(value):
    """Convert cell value to JSON-serializable format"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    # Handle basic types (str, int, float, bool)
    if isinstance(value, (str, int, float, bool)):
        return value
    # Convert any other object (ArrayFormula, etc.) to string
    return str(value)

def serialize_color(color):
    """Convert openpyxl color to JSON-serializable format"""
    if color is None:
        return None
    if hasattr(color, 'rgb'):
        rgb = color.rgb
        # RGB might be a string or RGB object
        return str(rgb) if rgb else None
    if hasattr(color, 'value'):
        return str(color.value) if color.value else None
    return str(color)

def extract_cell_data(cell):
    """Extract comprehensive cell data"""
    data = {
        "value": serialize_cell_value(cell.value),
        "data_type": cell.data_type,
        "coordinate": cell.coordinate,
    }

    # Formula
    if cell.data_type == 'f':
        # Handle both regular formulas and array formulas
        if cell.value:
            formula_str = str(cell.value)  # Convert ArrayFormula to string
            data["formula"] = f"={formula_str}"
        else:
            data["formula"] = None

    # Formatting
    if cell.font:
        data["font"] = {
            "name": cell.font.name,
            "size": cell.font.size,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "color": serialize_color(cell.font.color) if cell.font.color else None
        }

    if cell.fill:
        data["fill"] = {
            "pattern_type": cell.fill.patternType,
            "fg_color": serialize_color(cell.fill.fgColor) if cell.fill.fgColor else None,
            "bg_color": serialize_color(cell.fill.bgColor) if cell.fill.bgColor else None
        }

    # Number format
    if cell.number_format:
        data["number_format"] = cell.number_format

    # Hyperlink
    if cell.hyperlink:
        data["hyperlink"] = {
            "target": cell.hyperlink.target,
            "display": cell.hyperlink.display
        }

    # Comment
    if cell.comment:
        data["comment"] = {
            "text": cell.comment.text,
            "author": cell.comment.author
        }

    return data

def extract_workbook(xlsx_path, force=False):
    """Extract Excel workbook with full preservation"""
    xlsx_path = Path(xlsx_path)

    if not xlsx_path.exists():
        print(f"Error: Excel file not found: {xlsx_path}")
        return None

    # Initialize SmartCache (SHAKE256 hashing with auto-migration from SHA-256)
    smart_cache = SmartCache(doc_type='xlsx')

    # Generate cache key with SHAKE256 (auto-migrates from SHA-256 if needed)
    cache_key, cache_dir = smart_cache.get_cache_key(str(xlsx_path))
    cache_dir.mkdir(parents=True, exist_ok=True)

    # Check if already cached
    manifest_path = cache_dir / "manifest.json"
    if manifest_path.exists() and not force:
        print(f"Workbook already cached: {cache_key}")
        print(f"Cache location: {cache_dir}")
        print("Use --force to re-extract")
        return cache_key

    cache_dir.mkdir(exist_ok=True)

    print(f"Extracting Excel workbook: {xlsx_path.name}")
    print(f"File size: {xlsx_path.stat().st_size / 1024 / 1024:.2f} MB")

    # Open workbook
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False, keep_vba=False)
    except Exception as e:
        print(f"Error: Failed to open workbook: {e}")
        return None

    # Extract metadata
    metadata = {
        "creator": wb.properties.creator if wb.properties else None,
        "last_modified_by": wb.properties.lastModifiedBy if wb.properties else None,
        "created": wb.properties.created.isoformat() if wb.properties and wb.properties.created else None,
        "modified": wb.properties.modified.isoformat() if wb.properties and wb.properties.modified else None,
        "title": wb.properties.title if wb.properties else None,
        "subject": wb.properties.subject if wb.properties else None,
        "sheet_count": len(wb.sheetnames),
        "sheet_names": wb.sheetnames,
    }

    # Extract defined names (named ranges)
    named_ranges = {}
    if wb.defined_names:
        for name, definition in wb.defined_names.items():
            try:
                named_ranges[name] = {
                    "name": name,
                    "destinations": list(definition.destinations) if hasattr(definition, 'destinations') else []
                }
            except Exception:
                pass

    # Extract sheets
    sheets_data = []
    all_formulas = []

    print(f"Extracting {len(wb.sheetnames)} sheets...")

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        sheet = wb[sheet_name]

        print(f"  Processing sheet {sheet_idx + 1}/{len(wb.sheetnames)}: {sheet_name}")

        # Sheet metadata
        sheet_info = {
            "name": sheet_name,
            "sheet_state": sheet.sheet_state,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "dimensions": sheet.dimensions,
        }

        # Extract merged cells
        merged_cells = []
        if sheet.merged_cells:
            for merged_range in sheet.merged_cells.ranges:
                merged_cells.append(str(merged_range))
        sheet_info["merged_cells"] = merged_cells

        # Extract cells
        cells = []
        row_count = 0

        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                if cell.value is not None or cell.data_type == 'f':
                    cell_data = extract_cell_data(cell)
                    row_data.append(cell_data)

                    # Collect formulas
                    if cell.data_type == 'f' and cell.value:
                        all_formulas.append({
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "formula": f"={cell.value}"
                        })
                else:
                    row_data.append(None)

            cells.append(row_data)
            row_count += 1

            if row_count % 1000 == 0:
                print(f"    Processed {row_count}/{sheet.max_row} rows...")

        sheet_info["cells"] = cells
        sheet_info["row_count"] = row_count

        sheets_data.append(sheet_info)

        # Save individual sheet data
        sheets_dir = cache_dir / "sheets"
        sheets_dir.mkdir(exist_ok=True)

        sheet_filename = f"sheet_{sheet_idx + 1:03d}_{sheet_name.replace('/', '_')}.json"
        with open(sheets_dir / sheet_filename, "w", encoding="utf-8") as f:
            json.dump(sheet_info, f, indent=2, ensure_ascii=False)

    wb.close()

    # Save extracted data
    print("Saving extracted content...")

    # Full workbook data
    full_workbook = {
        "metadata": metadata,
        "sheets": sheets_data
    }

    with open(cache_dir / "full_workbook.json", "w", encoding="utf-8") as f:
        json.dump(full_workbook, f, indent=2, ensure_ascii=False)

    # Metadata
    with open(cache_dir / "metadata.json", "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2, ensure_ascii=False)

    # Named ranges
    with open(cache_dir / "named_ranges.json", "w", encoding="utf-8") as f:
        json.dump(named_ranges, f, indent=2, ensure_ascii=False)

    # Formulas
    with open(cache_dir / "formulas.json", "w", encoding="utf-8") as f:
        json.dump(all_formulas, f, indent=2, ensure_ascii=False)

    # Calculate statistics
    total_cells = sum(sheet["row_count"] * sheet["max_column"] for sheet in sheets_data)

    # Manifest
    manifest = {
        "cache_key": cache_key,
        "xlsx_name": xlsx_path.name,
        "xlsx_path": str(xlsx_path),
        "file_hash": cache_key.split('_')[-1],  # Extract hash from cache_key
        "sheet_count": metadata["sheet_count"],
        "total_cells": total_cells,
        "formula_count": len(all_formulas),
        "named_range_count": len(named_ranges),
        "cache_dir": str(cache_dir),
        "extracted_at": datetime.now().isoformat()
    }

    with open(cache_dir / "manifest.json", "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print(f"\nExtraction complete!")
    print(f"Cache key: {cache_key}")
    print(f"Cache location: {cache_dir}")
    print(f"Total sheets: {metadata['sheet_count']}")
    print(f"Total cells: {total_cells:,}")
    print(f"Formulas: {len(all_formulas):,}")
    print(f"Named ranges: {len(named_ranges)}")

    return cache_key

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_xlsx.py <xlsx_path> [--force] [--output-dir DIR]")
        print("\nOptions:")
        print("  --force          Force re-extraction even if cached")
        print("  --output-dir DIR Copy extracted files to specified directory (prompts if not provided)")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    force = "--force" in sys.argv

    # Extract output directory from arguments if provided
    output_dir = None
    if '--output-dir' in sys.argv:
        output_dir_index = sys.argv.index('--output-dir')
        if output_dir_index + 1 < len(sys.argv):
            output_dir = sys.argv[output_dir_index + 1]
        else:
            print("Error: --output-dir requires a value")
            sys.exit(1)

    cache_key = extract_workbook(xlsx_path, force)

    if cache_key:
        print(f"\nNext steps:")
        print(f"1. Chunk content: python scripts/chunk_sheets.py {cache_key}")
        print(f"2. Query content: python scripts/query_xlsx.py search {cache_key} 'your query'")

        # Handle output directory
        smart_cache = SmartCache(doc_type='xlsx')
        _, cache_path = smart_cache.get_cache_key(xlsx_path)
        should_copy = False
        keep_cache = True  # Default: always keep cache

        if output_dir is None:
            # Ask user interactively if they want to copy to working directory
            print("\n📁 Copy extracted files to working directory?")
            response = input("Copy files? (y/n): ").strip().lower()

            if response == 'y':
                output_dir = Path.cwd() / f"extracted_{cache_key}"
                should_copy = True

                # Ask about cache behavior only if copying
                print(f"\n💾 Maintain cache in {cache_path}?")
                print("  (yes) Keep cache for instant reuse across projects")
                print("  (no)  Remove cache after copying files")
                keep_cache_response = input("Keep cache? (y/n): ").strip().lower()
                keep_cache = keep_cache_response == 'y'
        else:
            # --output-dir was specified, skip prompts and copy files
            should_copy = True
            keep_cache = True  # Always keep cache when using --output-dir

        if should_copy and output_dir:
            output_path = Path(output_dir)

            # Copy all files from cache to output directory
            import shutil

            print(f"\n📋 Copying files to {output_path}...")
            output_path.mkdir(parents=True, exist_ok=True)

            # Copy all files and directories
            for item in cache_path.iterdir():
                if item.is_file():
                    shutil.copy2(item, output_path / item.name)
                    print(f"  ✓ {item.name}")
                elif item.is_dir():
                    shutil.copytree(item, output_path / item.name, dirs_exist_ok=True)
                    print(f"  ✓ {item.name}/ (directory)")

            print(f"\n✓ Files copied to: {output_path}")

            # Remove cache if requested
            if not keep_cache:
                shutil.rmtree(cache_path)
                print(f"  ✓ Cache removed")
            else:
                print(f"  ✓ Cache maintained at: {cache_path}")
                print(f"  💡 Cache persists across projects for instant reuse")
